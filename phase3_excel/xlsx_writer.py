from __future__ import annotations

from typing import Any, Dict, List, Sequence

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as e:  # pragma: no cover
    raise ModuleNotFoundError("缺少依赖 openpyxl，请先安装：pip install openpyxl") from e


def _set_column_widths(ws, headers: Sequence[str], rows: List[Dict[str, Any]]) -> None:
    # 按前 N 行做一个简单估算，避免 O(n*m) 的宽度计算太慢。
    sample = rows[:200]
    for i, h in enumerate(headers, start=1):
        values = [str(h)]
        for r in sample:
            v = r.get(h, "")
            values.append("" if v is None else str(v))
        max_len = max((len(v) for v in values), default=10)
        # 宽度给一个上限，避免超宽。
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)


def _merge_same_values_in_column(ws, *, col: int, start_row: int, end_row: int) -> None:
    """把同一列中连续相同的值合并成一个单元格（用于“章节/剧情块”列）。"""
    if end_row < start_row:
        return
    cur_start = start_row
    cur_val = ws.cell(row=start_row, column=col).value
    for r in range(start_row + 1, end_row + 1):
        v = ws.cell(row=r, column=col).value
        if v != cur_val:
            if cur_val not in (None, "") and r - 1 > cur_start:
                ws.merge_cells(start_row=cur_start, start_column=col, end_row=r - 1, end_column=col)
            cur_start = r
            cur_val = v
    if cur_val not in (None, "") and end_row > cur_start:
        ws.merge_cells(start_row=cur_start, start_column=col, end_row=end_row, end_column=col)


def _append_ai_footer(ws, *, columns: int, blank_rows: int = 3) -> None:
    """Append the AI attribution after some blank rows (merged + centered + bold)."""
    for _ in range(max(0, int(blank_rows))):
        ws.append(["" ] * columns)

    footer_row = ws.max_row + 1
    ws.append(["\u5185\u5bb9\u7531 AI \u4ea7\u751f"] + [""] * (columns - 1))
    if columns > 1:
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=columns)
    cell = ws.cell(row=footer_row, column=1)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_chunk_summary_merges(ws) -> None:
    """
    对 chunk 总结行做“临时合并”：
    - C:D 合并，显示“剧情概述/节奏概述”
    - E:G 合并，显示对应概述内容
    """
    label_plot = "\u5267\u60c5\u6982\u8ff0"
    label_pace = "\u8282\u594f\u6982\u8ff0"
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # 汇总行内容垂直居中（“行居中”），并保持左对齐便于阅读
    wrap_top = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Skip header row and footer row (footer row is appended later anyway).
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=3).value
        if v not in (label_plot, label_pace):
            continue

        # Merge "起始段落"+"结束段落" columns for the label.
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        c_label = ws.cell(row=r, column=3)
        c_label.font = bold
        c_label.alignment = center

        # Merge the last 3 columns (内容概述/节奏分析/爆点提取) for the summary text.
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        c_text = ws.cell(row=r, column=5)
        c_text.alignment = wrap_top
        # Excel won't auto-fit row height for merged cells; set a larger height so wrapping is visible.
        txt = c_text.value
        if txt not in (None, ""):
            ln = len(str(txt))
            lines = max(2, (ln + 79) // 80)  # rough estimate
            ws.row_dimensions[r].height = min(180, 18 * lines)



def _append_rows_with_chunk_gaps(
    ws,
    *,
    headers: Sequence[str],
    rows: List[Dict[str, Any]],
    chunk_gap_rows: int = 0,
    chapter_gap_rows: int = 1,
) -> None:
    """Append rows and insert blank separator rows between chapters/chunks."""
    if not headers:
        return

    chapter_key = headers[0]
    chunk_key = headers[1] if len(headers) > 1 else None

    last_chapter = None
    last_chunk = None
    for r in rows:
        row_values = [r.get(h, "") for h in headers]
        chapter_val = r.get(chapter_key, "")
        chunk_val = r.get(chunk_key, "") if chunk_key else ""

        cur_chunk = (chapter_val, chunk_val)

        # Only insert gaps when we have a meaningful chapter/chunk key.
        if chunk_val not in (None, ""):
            # Chapter boundary: add one blank row after finishing previous chapter.
            if last_chapter is not None and chapter_val != last_chapter:
                for _ in range(max(0, int(chapter_gap_rows))):
                    ws.append([""] * len(headers))
            # Chunk boundary (within the same chapter).
            elif last_chunk is not None and cur_chunk != last_chunk:
                for _ in range(max(0, int(chunk_gap_rows))):
                    ws.append([""] * len(headers))

            last_chapter = chapter_val
            last_chunk = cur_chunk

        ws.append(row_values)


def _apply_thin_grid_borders(ws, *, start_row: int, end_row: int, skip_blank_rows: bool = True) -> None:
    """Apply medium black borders to cells for readability."""
    side = Side(style='medium', color='808080')
    border = Border(left=side, right=side, top=side, bottom=side)

    for r in range(start_row, end_row + 1):
        if skip_blank_rows and r >= 2:
            # If the entire row is blank, keep it borderless as a visual separator.
            is_blank = True
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v not in (None, ""):
                    is_blank = False
                    break
            if is_blank:
                continue

        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border




def _apply_chapter_fill(ws, *, start_row: int, end_row: int) -> None:
    """Highlight chapter cells with a yellow background."""
    fill = PatternFill(fill_type='solid', fgColor='FFF2CC')
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=1)
        if cell.value not in (None, ''):
            cell.fill = fill
def write_table_sheet(
    wb: Workbook,
    *,
    title: str,
    headers: Sequence[str],
    rows: List[Dict[str, Any]],
    merge_same_value_cols: Sequence[int] = (),
    center_cols: Sequence[int] = (),
    bold_center_cols: Sequence[int] = (),
) -> None:
    ws = wb.create_sheet(title=title)

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header row
    ws.append(list(headers))
    ws.row_dimensions[1].height = 28
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_align

    # Data rows
    _append_rows_with_chunk_gaps(ws, headers=headers, rows=rows, chunk_gap_rows=0, chapter_gap_rows=1)

    # Formatting
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Wrap all cells (simple + robust)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = wrap

    # Merge repeated values (based on written values)
    data_end_row = ws.max_row
    for col in merge_same_value_cols:
        _merge_same_values_in_column(ws, col=col, start_row=2, end_row=data_end_row)

    # Column alignment styles
    for col in center_cols:
        for r in range(2, data_end_row + 1):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    bold_center = Font(bold=True)
    for col in bold_center_cols:
        for r in range(2, data_end_row + 1):
            c = ws.cell(row=r, column=col)
            c.font = bold_center
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _apply_chapter_fill(ws, start_row=2, end_row=data_end_row)
    _apply_chunk_summary_merges(ws)
    _append_ai_footer(ws, columns=len(headers))
    _apply_thin_grid_borders(ws, start_row=1, end_row=ws.max_row, skip_blank_rows=True)
    _set_column_widths(ws, headers, rows)


def build_workbook(
    *,
    rows: List[Dict[str, Any]],
) -> Workbook:
    wb = Workbook()
    # Remove the default sheet to avoid confusion.
    wb.remove(wb.active)

    write_table_sheet(
        wb,
        title="分析",
        headers=[
            "章节",
            "剧情块",
            "起始段落",
            "结束段落",
            "内容概述",
            "节奏分析",
            "爆点提取",
        ],
        rows=rows,
        merge_same_value_cols=(1, 2),
        center_cols=(3, 4),
        bold_center_cols=(1, 2),
    )

    return wb
